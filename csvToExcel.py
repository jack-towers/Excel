import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from openpyxl.cell.cell import MergedCell
def process_service_status(csv_path, output_path="Formatted_Service_Status.xlsx"):
    # Import dependencies
    import pandas as pd
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.cell.cell import MergedCell
    import re

    # Load CSV
    df = pd.read_csv(csv_path)

    # Remove the first column (Location)
    df = df.iloc[:, 1:]

    # Keep only required columns in order
    columns_to_keep = [
        "Name", "Model", "Serial Number",
        "Date of Last Service", "Last Service Description",
        "Next Service Description", "Next Service Due", "Unspc Code"
    ]
    df = df[columns_to_keep]

    # Remove offline assets
    df = df[~df.apply(lambda row: row.astype(str).str.contains("offline", case=False, na=False).any(), axis=1)]

    # Convert date columns
    date_columns = ["Date of Last Service", "Next Service Due"]
    for col in date_columns:
        df[col] = pd.to_datetime(df[col], format="%b %d, %Y", errors='coerce')

    # Remove rows where all service-related fields are blank
    df = df.dropna(
        subset=["Date of Last Service", "Last Service Description", "Next Service Description", "Next Service Due"],
        how='all')

    # Calculate weeks overdue (calendar weeks)
    today = datetime.today()
    df["Weeks Overdue for Service"] = ((today - df["Next Service Due"]).dt.days // 7).fillna(0).astype(int)

    # Group model names to remove everything after "-" or "/"
    df["Cleaned Model"] = df["Model"].apply(lambda x: re.split(r"[-/]", str(x))[0])

    # Sort by UNSPSC Code first, then Cleaned Model
    df = df.sort_values(by=["Unspc Code", "Cleaned Model"])

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Service Status"

    # Define styles
    family_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    family_header_font = Font(size=16, bold=True, color="FFFFFF")
    family_header_alignment = Alignment(horizontal="left", vertical="center")

    column_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    column_header_font = Font(bold=True, color="FFFFFF")
    column_header_alignment = Alignment(horizontal="center", vertical="center")

    subheader_font = Font(size=12, bold=True)
    subheader_alignment = Alignment(horizontal="left", vertical="center")

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    # Group by UNSPSC Code first, then by Cleaned Model
    unspsc_groups = df.groupby("Unspc Code")
    row_idx = 1  # Start from the first row

    for unspsc_code, unspsc_group in unspsc_groups:
        # Write UNSPSC Code as a major family heading
        ws.append([unspsc_code])  # Write UNSPSC Code
        family_cell = ws.cell(row=row_idx, column=1)
        family_cell.font = family_header_font
        family_cell.fill = family_header_fill
        family_cell.alignment = family_header_alignment
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(columns_to_keep))
        row_idx += 1

        model_groups = unspsc_group.groupby("Cleaned Model")
        for model, group in model_groups:
            # Write Model as a subsection heading
            ws.append([model])  # Write Model
            subheader_cell = ws.cell(row=row_idx, column=1)
            subheader_cell.font = subheader_font
            subheader_cell.alignment = subheader_alignment
            row_idx += 1

            # Write the actual header row for the data
            for r_idx, row in enumerate(dataframe_to_rows(group.drop(columns=["Cleaned Model", "Unspc Code"]),
                                                          index=False, header=True), start=row_idx):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # Format column header row
                    if r_idx == row_idx:
                        cell.fill = column_header_fill
                        cell.font = column_header_font
                        cell.alignment = column_header_alignment

            # Update row_idx for the next model
            row_idx += len(group) + 1

    # Apply formatting based on overdue status
    weeks_overdue_col_idx = df.columns.get_loc("Weeks Overdue for Service")
    print(f"Weeks Overdue column index: {weeks_overdue_col_idx}")

    # Now loop through the rows to apply formatting
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=weeks_overdue_col_idx,
                            max_col=weeks_overdue_col_idx):
        for cell in row:
            print(f"Processing cell: {cell.coordinate}, value: {cell.value}")
            if isinstance(cell.value, (int, float)):
                if cell.value > 25:
                    cell.fill = red_fill
                elif 0 <= cell.value <= 25:
                    cell.fill = yellow_fill
                else:
                    cell.fill = green_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell) or cell.value is None:  # Skip merged/empty cells
                continue
            max_length = max(max_length, len(str(cell.value)))
            if not col_letter:
                col_letter = cell.column_letter
        if col_letter:
            ws.column_dimensions[col_letter].width = max_length + 2

    # Save the formatted Excel file
    wb.save(output_path)
    print(f"Processed file saved as: {output_path}")



# Example usage:
# process_service_status("your_csv_file.csv")
process_service_status("Assets (63).csv")
